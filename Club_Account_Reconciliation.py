# -*- coding: utf-8 -*-
"""
Created on Wed Jul 10 16:42:39 2019

Author: Richie Youm
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook

#file path for the excel file from RBC Express
file_path="C:/Users/vbashkatova/Desktop/Bank_reconciliation/Jan 31, 20.xlsx"

#where the generated report would be added to
report_path='../SSMU Monthly Bank Reconciliation [2018-2019].xlsx'

def preprocessing(file_path):
    df=pd.read_excel(file_path).iloc[1:,:].reset_index(drop=True)
    df=df.iloc[:,[0,1,3,4]]
    df.columns=['Account Type','Account','Currency','Balance']
    # dropping last "End of report" line
    df=df.dropna()
    # cleaning 'Account' column
    df['Account']=[i.replace('ROYAL BANK OF\nCANADA-','') for i in df['Account']] 
    # cleaing CLU 1 & 2 (previously called SSMU-1 and SSMU-2)
    df['Account']=[i.replace('ROYAL BANK OF CANADA-','') for i in df['Account']] 
    df['Account number']=[i[1] for i in df['Account'].str.split('-',n=1)]
    df['Account']=df['Account'].str.split('-',n=1,expand=True)
    df=df.reset_index(drop=True)
    df=df.set_index('Account Type')
    df=df[['Account','Account number', 'Currency', 'Balance']]
    return df

def generate(df, report_path, prev_mo_end, reporting_end_date):
    """ Generate a new report sheet and appends it to the main report file
    Parameters:
        df (DataFrame): preprocessed RBC data
        report_path (String): path of the main report file
        prev_mo_end (String): previous month in the report. Case sensitive
        reporting_end_date (String): name to be used for current report sheet. Case sensitive
    """
    book=load_workbook(report_path)
    writer=pd.ExcelWriter(report_path, engine='openpyxl')
    writer.book=book
    df.to_excel(writer,sheet_name=reporting_end_date, startrow=12)
    # writer declaration
    ws=writer.sheets[reporting_end_date]
    ws_prev=book[prev_mo_end]
    # specific content position
    ws['A3']='Balance Reporting - Balance Summary Report'
    ws['D{}'.format(df.shape[0]+16)]='Grand Total'
    ws['E{}'.format(df.shape[0]+16)]=np.sum(df['Balance'])
    ws['C6']='Monthly Total Changes'
    ws['E5']='Adjustments'
    ws.merge_cells('C6:D6')
    ws.merge_cells('E5:F5')
    ws['C7']=reporting_end_date
    ws['C8']='Previous month end'
    ws['C9']='Monthly change'
    ws['D7']=np.sum(df['Balance'])
    ws['D8']=ws_prev['D7'].value
    ws['D9']=ws['D7'].value-ws['D8'].value
    ws['E6']='SSMU-1'
    ws['E7']=df[df['Account']=='CLU1']['Balance'][0]
    ws['E8']=ws_prev['E7'].value
    ws['F6']='SSMU-2'
    ws['F7']=df[df['Account']=='CLU2']['Balance'][0]
    ws['F8']=ws_prev['F7'].value
    ws['G6']='Net Total'
    ws['G7']=ws['D7'].value-ws['E7'].value-ws['F7'].value
    ws['G8']=ws['D8'].value-ws['E8'].value-ws['F8'].value
    ws['G9']=ws['G7'].value-ws['G8'].value
    writer.save()
    print('All done! Check the excel file.')

def main():
    prev_mo_end=input('Please enter the sheet name of previous month end: \n')
    reporting_end_date=input('Please enter the sheet name of current month end: \n')
    df = preprocessing(file_path)
    generate(df, report_path, prev_mo_end, reporting_end_date)

main()